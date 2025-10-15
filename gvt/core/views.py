from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import logout, login as auth_login
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Avg, Q
from django.utils import timezone
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from datetime import timedelta
import json
import io
import os

from .models import (
    CybercrimeReport, EvidenceFile, CRIME_TYPES, 
    ActivityLog, SystemMetrics, UserSession, ReportComment
)
from .forms import CybercrimeReportForm
from .EmailBackEnd import EmailBackEnd

# Import for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, ListFlowable, ListItem, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import qrcode

# Import for Excel export
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ============= HELPER FUNCTIONS =============

def log_activity(user, action, content_object=None, description=None, old_value=None, new_value=None, request=None):
    """Universal activity logging function"""
    log_data = {
        'user': user if user and user.is_authenticated else None,
        'action': action,
        'description': description,
        'old_value': old_value,
        'new_value': new_value,
    }
    
    if content_object:
        log_data['content_type'] = ContentType.objects.get_for_model(content_object)
        log_data['object_id'] = content_object.id
    
    if request:
        log_data['ip_address'] = get_client_ip(request)
        log_data['user_agent'] = request.META.get('HTTP_USER_AGENT', '')[:500]
        log_data['session_key'] = request.session.session_key
    
    ActivityLog.objects.create(**log_data)


def get_client_ip(request):
    """Extract client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def update_system_metrics():
    """Update daily system metrics"""
    today = timezone.now().date()
    metrics, created = SystemMetrics.objects.get_or_create(date=today)
    
    reports = CybercrimeReport.objects.filter(submitted_at__date=today)
    
    metrics.total_reports = CybercrimeReport.objects.count()
    metrics.reports_created = reports.count()
    metrics.reports_resolved = reports.filter(is_resolved=True).count()
    metrics.reports_updated = ActivityLog.objects.filter(
        action='update',
        timestamp__date=today
    ).count()
    
    avg_priority = CybercrimeReport.objects.aggregate(Avg('priority_score'))
    metrics.average_priority_score = avg_priority['priority_score__avg'] or 0.0
    
    metrics.critical_reports = CybercrimeReport.objects.filter(priority_level='critical').count()
    metrics.high_priority_reports = CybercrimeReport.objects.filter(priority_level='high').count()
    metrics.medium_priority_reports = CybercrimeReport.objects.filter(priority_level='medium').count()
    metrics.low_priority_reports = CybercrimeReport.objects.filter(priority_level='low').count()
    
    metrics.unique_users = ActivityLog.objects.filter(
        timestamp__date=today
    ).values('user').distinct().count()
    
    metrics.total_evidence_files = EvidenceFile.objects.count()
    
    metrics.save()


# ============= AUTHENTICATION VIEWS =============

def login_view(request):
    if request.method == "POST":
        user = EmailBackEnd.authenticate(
            request,
            username=request.POST.get('email'),
            password=request.POST.get('password')
        )
        if user is not None:
            auth_login(request, user)
            
            # Create user session
            UserSession.objects.create(
                user=user,
                session_key=request.session.session_key,
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
            )
            
            # Log login activity
            log_activity(
                user=user,
                action='login',
                description=f"User {user.username} logged in",
                request=request
            )
            
            messages.success(request, f"Welcome back, {user.username}!")
            return redirect('dashboard')
        else:
            log_activity(
                user=None,
                action='login',
                description=f"Failed login attempt for email: {request.POST.get('email')}",
                request=request
            )
            messages.error(request, "Invalid Login Credentials!")
            return render(request, 'login.html')
    else:
        return render(request, 'login.html')


@login_required
def logout_view(request):
    # Update user session
    try:
        session = UserSession.objects.get(
            user=request.user,
            session_key=request.session.session_key,
            is_active=True
        )
        session.logout_time = timezone.now()
        session.is_active = False
        session.save()
    except UserSession.DoesNotExist:
        pass
    
    # Log logout activity
    log_activity(
        user=request.user,
        action='logout',
        description=f"User {request.user.username} logged out",
        request=request
    )
    
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('report_cybercrime')


# ============= REPORT CRUD VIEWS =============

def report_cybercrime(request):
    """Create new cybercrime report"""
    if request.method == 'POST':
        form = CybercrimeReportForm(request.POST, request.FILES)
        if form.is_valid():
            report = form.save(commit=False)
            
            # Set optional fields
            report.latitude = request.POST.get('latitude', '')
            report.longitude = request.POST.get('longitude', '')
            report.browser_info = request.POST.get('browser_info', '')
            report.device_info = request.POST.get('device_info', '')
            report.ip_address = get_client_ip(request)
            
            if request.user.is_authenticated:
                report.created_by = request.user
            
            report.save()
            
            # Handle evidence files
            files = request.FILES.getlist('evidence_files')
            for file in files:
                evidence = EvidenceFile.objects.create(
                    report=report,
                    file=file,
                    uploaded_by=request.user if request.user.is_authenticated else None
                )
                log_activity(
                    user=request.user if request.user.is_authenticated else None,
                    action='upload',
                    content_object=evidence,
                    description=f"Evidence file '{file.name}' uploaded for report #{report.id}",
                    request=request
                )
            
            # Calculate priority score
            report.calculate_priority_score()
            
            # Log report creation
            log_activity(
                user=request.user if request.user.is_authenticated else None,
                action='create',
                content_object=report,
                description=f"New {report.get_crime_type_display()} report created",
                new_value={
                    'crime_type': report.crime_type,
                    'priority_level': report.priority_level,
                    'priority_score': report.priority_score,
                },
                request=request
            )
            
            # Update system metrics
            update_system_metrics()
            
            messages.success(request, 'Report submitted successfully!')
            return redirect('success', report_id=report.id)
    else:
        form = CybercrimeReportForm()

    return render(request, 'report_cybercrime.html', {'form': form})


def report_success(request, report_id):
    """Display success page after report submission"""
    report = get_object_or_404(CybercrimeReport, id=report_id)
    evidence_files = report.evidence_files.all()
    
    # Log view activity (only for authenticated users to avoid spam)
    if request.user.is_authenticated:
        log_activity(
            user=request.user,
            action='view',
            content_object=report,
            description=f"Viewed success page for report #{report.id}",
            request=request
        )
    
    return render(request, 'report_success.html', {
        'message': 'Report submitted successfully!',
        'report': report,
        'evidence_files': evidence_files
    })


@login_required
def edit_report(request, report_id):
    """Edit existing cybercrime report"""
    report = get_object_or_404(CybercrimeReport, id=report_id)
    
    # Store old values for logging
    old_values = {
        'crime_type': report.crime_type,
        'description': report.description,
        'priority_level': report.priority_level,
        'incident_date': str(report.incident_date),
    }
    
    if request.method == 'POST':
        form = CybercrimeReportForm(request.POST, request.FILES, instance=report)
        if form.is_valid():
            report = form.save(commit=False)
            report.updated_by = request.user
            report.save()
            
            # Recalculate priority if needed
            if form.has_changed():
                report.calculate_priority_score()
            
            # Handle new evidence files
            files = request.FILES.getlist('evidence_files')
            for file in files:
                evidence = EvidenceFile.objects.create(
                    report=report,
                    file=file,
                    uploaded_by=request.user
                )
                log_activity(
                    user=request.user,
                    action='upload',
                    content_object=evidence,
                    description=f"Additional evidence '{file.name}' uploaded",
                    request=request
                )
            
            # Store new values
            new_values = {
                'crime_type': report.crime_type,
                'description': report.description,
                'priority_level': report.priority_level,
                'incident_date': str(report.incident_date),
            }
            
            # Log update activity
            log_activity(
                user=request.user,
                action='update',
                content_object=report,
                description=f"Report #{report.id} updated",
                old_value=old_values,
                new_value=new_values,
                request=request
            )
            
            update_system_metrics()
            messages.success(request, 'Report updated successfully!')
            return redirect('view_report_analysis', report_id=report.id)
    else:
        form = CybercrimeReportForm(instance=report)
    
    # Log view activity
    log_activity(
        user=request.user,
        action='view',
        content_object=report,
        description=f"Opened edit form for report #{report.id}",
        request=request
    )
    
    return render(request, 'edit_report.html', {
        'form': form,
        'report': report,
        'evidence_files': report.evidence_files.all()
    })


@login_required
def delete_report(request, report_id):
    """Delete cybercrime report"""
    report = get_object_or_404(CybercrimeReport, id=report_id)
    
    if request.method == 'POST':
        # Store report info before deletion
        report_info = {
            'id': report.id,
            'crime_type': report.crime_type,
            'description': report.description[:100],
            'priority_level': report.priority_level,
        }
        
        # Log deletion
        log_activity(
            user=request.user,
            action='delete',
            description=f"Deleted report #{report.id} - {report.get_crime_type_display()}",
            old_value=report_info,
            request=request
        )
        
        report.delete()
        update_system_metrics()
        messages.success(request, 'Report deleted successfully!')
        return redirect('dashboard')
    
    # Log view of delete confirmation
    log_activity(
        user=request.user,
        action='view',
        content_object=report,
        description=f"Viewed delete confirmation for report #{report.id}",
        request=request
    )
    
    return render(request, 'confirm_delete.html', {'report': report})


# ============= DASHBOARD AND ANALYSIS VIEWS =============
@login_required
def dashboard(request):
    """Main dashboard view"""
    reports = CybercrimeReport.objects.all().order_by('-submitted_at')
    
    # Calculate statistics
    resolved_reports = reports.filter(is_resolved=True).count()
    pending_reports = reports.filter(is_resolved=False).count()
    critical_reports = reports.filter(priority_level='critical').count()
    high_reports = reports.filter(priority_level='high').count()
    medium_reports = reports.filter(priority_level='medium').count()
    low_reports = reports.filter(priority_level='low').count()
    
    # Active users (last 24 hours)
    active_users = ActivityLog.objects.filter(
        timestamp__gte=timezone.now() - timedelta(hours=24)
    ).values('user').distinct().count()
    
    # Chart data - crime type distribution
    crime_type_data = []
    crime_type_labels = []
    for crime_type, display_name in CRIME_TYPES:
        count = reports.filter(crime_type=crime_type).count()
        if count > 0:
            crime_type_data.append(count)
            crime_type_labels.append(display_name)
    
    # Priority distribution for chart
    priority_data = [critical_reports, high_reports, medium_reports, low_reports]
    priority_labels = ['Critical', 'High', 'Medium', 'Low']
    
    # Recent activities
    recent_activities = ActivityLog.objects.all().order_by('-timestamp')[:10]
    
    # Log dashboard access
    log_activity(
        user=request.user,
        action='view',
        description="Accessed main dashboard",
        request=request
    )
    
    context = {
        'reports': reports,
        'CRIME_TYPES': CRIME_TYPES,
        'resolved_reports': resolved_reports,
        'pending_reports': pending_reports,
        'critical_reports': critical_reports,
        'high_reports': high_reports,
        'medium_reports': medium_reports,
        'low_reports': low_reports,
        'active_users': active_users,
        'recent_activities': recent_activities,
        'crime_type_data': crime_type_data,
        'crime_type_labels': crime_type_labels,
        'priority_data': priority_data,
        'priority_labels': priority_labels,
    }
    return render(request, 'dashboard.html', context)

@login_required
def priority_dashboard(request):
    """Dashboard showing reports categorized by priority levels"""
    priority_filter = request.GET.get('priority', '')
    date_filter = request.GET.get('date_range', '')
    
    reports = CybercrimeReport.objects.all()
    
    if priority_filter:
        reports = reports.filter(priority_level=priority_filter)
        
    if date_filter:
        if date_filter == 'today':
            reports = reports.filter(submitted_at__date=timezone.now().date())
        elif date_filter == 'week':
            one_week_ago = timezone.now() - timedelta(days=7)
            reports = reports.filter(submitted_at__gte=one_week_ago)
        elif date_filter == 'month':
            one_month_ago = timezone.now() - timedelta(days=30)
            reports = reports.filter(submitted_at__gte=one_month_ago)
    
    priority_stats = reports.values('priority_level').annotate(count=Count('id'))
    crime_type_stats = reports.values('crime_type').annotate(count=Count('id'))
    
    log_activity(
        user=request.user,
        action='view',
        description=f"Accessed priority dashboard (filters: priority={priority_filter}, date={date_filter})",
        request=request
    )
    
    context = {
        'reports': reports.order_by('-priority_score', '-submitted_at'),
        'priority_stats': priority_stats,
        'crime_type_stats': crime_type_stats,
        'selected_priority': priority_filter,
        'selected_date_range': date_filter,
    }
    return render(request, 'priority_dashboard.html', context)


@login_required
def view_report_analysis(request, report_id):
    """Detailed AI analysis view for a specific report"""
    report = get_object_or_404(CybercrimeReport, id=report_id)
    
    similar_reports = CybercrimeReport.objects.filter(
        crime_type=report.crime_type
    ).exclude(id=report.id).order_by('-priority_score')[:5]
    
    trend_data = CybercrimeReport.objects.filter(
        crime_type=report.crime_type,
        submitted_at__gte=timezone.now() - timedelta(days=30)
    ).values('submitted_at__date').annotate(
        count=Count('id'),
        avg_priority=Avg('priority_score')
    ).order_by('submitted_at__date')
    
    # Get comments
    comments = report.comments.all()
    
    # Get activity log for this report
    content_type = ContentType.objects.get_for_model(CybercrimeReport)
    activities = ActivityLog.objects.filter(
        content_type=content_type,
        object_id=report.id
    )[:20]
    
    log_activity(
        user=request.user,
        action='view',
        content_object=report,
        description=f"Viewed detailed analysis for report #{report.id}",
        request=request
    )
    
    context = {
        'report': report,
        'similar_reports': similar_reports,
        'trend_data': list(trend_data),
        'ai_analysis': report.ai_analysis,
        'comments': comments,
        'activities': activities,
        'evidence_files': report.evidence_files.all(),
    }
    return render(request, 'report_analysis.html', context)



@login_required
def update_report_status(request, report_id):
    """Update the status of a cybercrime report"""
    report = get_object_or_404(CybercrimeReport, id=report_id)

    if request.method == 'POST':
        # Store old status for logging
        old_status = report.is_resolved

        # Get new status from form
        status = request.POST.get('status') == 'True'  # Convert string 'True'/'False' to boolean

        # Update report status
        report.is_resolved = status
        if status:
            report.resolution_date = timezone.now()
            report.resolved_by = request.user
        else:
            report.resolution_date = None
            report.resolved_by = None
        report.save()

        # Log status update
        log_activity(
            user=request.user,
            action='update',
            content_object=report,
            description=f"Updated status of report #{report.id} from {'Resolved' if old_status else 'Unresolved'} to {'Resolved' if status else 'Unresolved'}",
            old_value={'is_resolved': old_status},
            new_value={'is_resolved': status},
            request=request
        )

        # Update system metrics
        update_system_metrics()

        messages.success(request, f"Report status updated to {'Resolved' if status else 'Unresolved'}.")
        return redirect('view_report_analysis', report_id=report.id)

    # If GET, show a confirmation page (optional, but redirecting to analysis page for simplicity)
    messages.error(request, "Invalid request method for updating status.")
    return redirect('view_report_analysis', report_id=report.id)

# ============= COMMENT VIEWS =============

@login_required
def add_comment(request, report_id):
    """Add comment to a report"""
    if request.method == 'POST':
        report = get_object_or_404(CybercrimeReport, id=report_id)
        comment_text = request.POST.get('comment', '').strip()
        is_internal = request.POST.get('is_internal', 'true') == 'true'
        
        if comment_text:
            comment = ReportComment.objects.create(
                report=report,
                user=request.user,
                comment=comment_text,
                is_internal=is_internal
            )
            
            log_activity(
                user=request.user,
                action='create',
                content_object=comment,
                description=f"Added {'internal' if is_internal else 'public'} comment to report #{report.id}",
                new_value={'comment': comment_text[:100]},
                request=request
            )
            
            messages.success(request, 'Comment added successfully!')
        else:
            messages.error(request, 'Comment cannot be empty.')
        
        return redirect('view_report_analysis', report_id=report.id)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


@login_required
def edit_comment(request, comment_id):
    """Edit an existing comment"""
    if request.method == 'POST':
        comment = get_object_or_404(ReportComment, id=comment_id, user=request.user)
        old_text = comment.comment
        new_text = request.POST.get('comment', '').strip()
        
        if new_text:
            comment.comment = new_text
            comment.save()
            
            log_activity(
                user=request.user,
                action='update',
                content_object=comment,
                description=f"Edited comment on report #{comment.report.id}",
                old_value={'comment': old_text[:100]},
                new_value={'comment': new_text[:100]},
                request=request
            )
            
            messages.success(request, 'Comment updated successfully!')
        else:
            messages.error(request, 'Comment cannot be empty.')
        
        return redirect('view_report_analysis', report_id=comment.report.id)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


@login_required
def delete_comment(request, comment_id):
    """Delete a comment"""
    if request.method == 'POST':
        comment = get_object_or_404(ReportComment, id=comment_id, user=request.user)
        report_id = comment.report.id
        comment_text = comment.comment[:100]
        
        log_activity(
            user=request.user,
            action='delete',
            description=f"Deleted comment from report #{report_id}",
            old_value={'comment': comment_text},
            request=request
        )
        
        comment.delete()
        messages.success(request, 'Comment deleted successfully!')
        return redirect('view_report_analysis', report_id=report_id)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


# ============= EXPORT VIEWS =============

@login_required
def export_report(request, report_id):
    """Export report as PDF"""
    try:
        report = get_object_or_404(CybercrimeReport, id=report_id)
        evidence_files = report.evidence_files.all()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(f"Cybercrime Report #{report.id}", styles['Title']))
        elements.append(Spacer(1, 12))

        # QR Code
        site_url = getattr(settings, 'SITE_URL', request.build_absolute_uri('/')[:-1])
        qr_data = f"{site_url}/report/{report.id}/"
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_buffer = io.BytesIO()
        qr_img.save(qr_buffer, format="PNG")
        qr_buffer.seek(0)
        qr_image = Image(qr_buffer, width=1*inch, height=1*inch)
        elements.append(qr_image)
        elements.append(Spacer(1, 12))

        # Incident Details
        elements.append(Paragraph("Incident Details", styles['Heading2']))
        elements.append(Paragraph(f"Crime Type: {report.get_crime_type_display()}", styles['Normal']))
        elements.append(Paragraph(f"Priority: {report.get_priority_level_display()} ({report.priority_score:.1f})", styles['Normal']))
        elements.append(Paragraph(f"Status: {'Resolved' if report.is_resolved else 'Open'}", styles['Normal']))
        elements.append(Paragraph(f"Incident Date: {report.incident_date.strftime('%B %d, %Y, %I:%M %p')}", styles['Normal']))
        elements.append(Paragraph(f"Description: {report.description}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Reporter Information
        elements.append(Paragraph("Reporter Information", styles['Heading2']))
        elements.append(Paragraph(f"Name: {report.reporter_name or 'Anonymous'}", styles['Normal']))
        elements.append(Paragraph(f"Email: {report.reporter_email or 'Not provided'}", styles['Normal']))
        elements.append(Paragraph(f"Phone: {report.reporter_phone or 'Not provided'}", styles['Normal']))
        location = f"{report.latitude}, {report.longitude}" if report.latitude and report.longitude else "Not provided"
        elements.append(Paragraph(f"Location: {location}", styles['Normal']))
        elements.append(Paragraph(f"IP Address: {report.ip_address or 'Not captured'}", styles['Normal']))
        elements.append(Spacer(1, 12))

        # Evidence Files
        elements.append(Paragraph("Evidence Files", styles['Heading2']))
        if evidence_files:
            evidence_list = [ListItem(Paragraph(f"{evidence.file.name}", styles['Normal'])) for evidence in evidence_files]
            elements.append(ListFlowable(evidence_list, bulletType='bullet'))
        else:
            elements.append(Paragraph("No evidence files uploaded.", styles['Normal']))
        elements.append(Spacer(1, 12))

        # AI Analysis Summary
        if report.ai_analysis:
            elements.append(Paragraph("AI Analysis Summary", styles['Heading2']))
            elements.append(Paragraph(f"Total Priority Score: {report.ai_analysis.get('total_score', 0):.1f}/100", styles['Normal']))
            elements.append(Spacer(1, 6))

        doc.build(elements)
        buffer.seek(0)

        # Log export
        log_activity(
            user=request.user,
            action='export',
            content_object=report,
            description=f"Exported report #{report.id} as PDF",
            request=request
        )

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Report_{report.id}.pdf"'
        return response

    except CybercrimeReport.DoesNotExist:
        messages.error(request, "Report not found")
        return redirect('dashboard')


@login_required
def export_all_data_to_excel(request):
    """Export all data to Excel"""
    wb = Workbook()
    
    # Cybercrime Reports Sheet
    ws_reports = wb.active
    ws_reports.title = "Cybercrime Reports"
    
    report_headers = [
        "ID", "Crime Type", "Priority Level", "Priority Score", "Status",
        "Incident Date", "Description", "Reporter Name", "Reporter Email", 
        "Reporter Phone", "Additional Info", "Submitted At", "Updated At",
        "Latitude", "Longitude", "Browser Info", "IP Address", "Device Info",
        "Created By", "Updated By", "Resolved By", "Resolution Date"
    ]
    ws_reports.append(report_headers)
    
    reports = CybercrimeReport.objects.all()
    for report in reports:
        ws_reports.append([
            report.id,
            report.get_crime_type_display(),
            report.get_priority_level_display(),
            report.priority_score,
            'Resolved' if report.is_resolved else 'Open',
            report.incident_date.strftime('%Y-%m-%d %H:%M:%S') if report.incident_date else '',
            report.description,
            report.reporter_name or '',
            report.reporter_email or '',
            report.reporter_phone or '',
            report.additional_info or '',
            report.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if report.submitted_at else '',
            report.updated_at.strftime('%Y-%m-%d %H:%M:%S') if report.updated_at else '',
            report.latitude,
            report.longitude,
            report.browser_info or '',
            report.ip_address or '',
            report.device_info or '',
            report.created_by.username if report.created_by else '',
            report.updated_by.username if report.updated_by else '',
            report.resolved_by.username if report.resolved_by else '',
            report.resolution_date.strftime('%Y-%m-%d %H:%M:%S') if report.resolution_date else ''
        ])
    
    # Evidence Files Sheet
    ws_evidence = wb.create_sheet(title="Evidence Files")
    evidence_headers = ["Report ID", "File Name", "Uploaded At", "Uploaded By"]
    ws_evidence.append(evidence_headers)
    
    evidence_files = EvidenceFile.objects.all()
    for evidence in evidence_files:
        ws_evidence.append([
            evidence.report.id,
            evidence.file.name,
            evidence.uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if evidence.uploaded_at else '',
            evidence.uploaded_by.username if evidence.uploaded_by else 'Anonymous'
        ])
    
    # Activity Log Sheet
    ws_activity = wb.create_sheet(title="Activity Log")
    activity_headers = ["ID", "User", "Action", "Timestamp", "Description", "IP Address"]
    ws_activity.append(activity_headers)
    
    activities = ActivityLog.objects.all()[:1000]  # Last 1000 activities
    for activity in activities:
        ws_activity.append([
            activity.id,
            activity.user.username if activity.user else 'Anonymous',
            activity.get_action_display(),
            activity.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            activity.description or '',
            activity.ip_address or ''
        ])
    
    # System Metrics Sheet
    ws_metrics = wb.create_sheet(title="System Metrics")
    metrics_headers = [
        "Date", "Total Reports", "Reports Created", "Reports Resolved",
        "Avg Priority Score", "Critical", "High", "Medium", "Low", "Unique Users"
    ]
    ws_metrics.append(metrics_headers)
    
    metrics = SystemMetrics.objects.all()[:90]  # Last 90 days
    for metric in metrics:
        ws_metrics.append([
            metric.date.strftime('%Y-%m-%d'),
            metric.total_reports,
            metric.reports_created,
            metric.reports_resolved,
            metric.average_priority_score,
            metric.critical_reports,
            metric.high_priority_reports,
            metric.medium_priority_reports,
            metric.low_priority_reports,
            metric.unique_users
        ])
    
    # Adjust column widths
    for ws in [ws_reports, ws_evidence, ws_activity, ws_metrics]:
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Log export
    log_activity(
        user=request.user,
        action='export',
        description="Exported all system data to Excel",
        request=request
    )
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Cybercrime_Complete_Data.xlsx"'
    return response


# ============= STATISTICS AND ANALYTICS VIEWS =============

@login_required
def statistics_view(request):
    """View for showing overall statistics and trends"""
    now = timezone.now()
    last_month = now - timedelta(days=30)
    
    monthly_stats = (
        CybercrimeReport.objects
        .filter(submitted_at__gte=last_month)
        .values('submitted_at__date')
        .annotate(
            count=Count('id'),
            avg_priority=Avg('priority_score')
        )
        .order_by('submitted_at__date')
    )
    
    monthly_stats_list = [
        {
            'submitted_at__date': item['submitted_at__date'].strftime('%Y-%m-%d'),
            'count': item['count'],
            'avg_priority': float(item['avg_priority']) if item['avg_priority'] else 0
        }
        for item in monthly_stats
    ]
    
    crime_distribution = list(
        CybercrimeReport.objects
        .values('crime_type')
        .annotate(
            count=Count('id'),
            avg_priority=Avg('priority_score')
        )
    )
    
    crime_type_dict = dict(CRIME_TYPES)
    total_reports = sum(item['count'] for item in crime_distribution)
    colors = ['#1a73e8', '#ea4335', '#34a853', '#fbbc04', '#9334e8', '#ff6d01', '#46bdc6', '#7c4dff']
    
    for i, item in enumerate(crime_distribution):
        item['display_name'] = crime_type_dict.get(item['crime_type'], item['crime_type'])
        item['avg_priority'] = float(item['avg_priority']) if item['avg_priority'] else 0
        item['percentage'] = (item['count'] / total_reports * 100) if total_reports > 0 else 0
        item['color'] = colors[i % len(colors)]
    
    priority_distribution = (
        CybercrimeReport.objects
        .values('priority_level')
        .annotate(count=Count('id'))
        .order_by('-count')
    )
    
    # Get recent system metrics
    recent_metrics = SystemMetrics.objects.all()[:30]
    
    log_activity(
        user=request.user,
        action='view',
        description="Accessed statistics dashboard",
        request=request
    )
    
    context = {
        'monthly_stats': monthly_stats_list,
        'crime_distribution': crime_distribution,
        'priority_distribution': priority_distribution,
        'total_reports': CybercrimeReport.objects.count(),
        'recent_metrics': recent_metrics,
    }
    return render(request, 'statistics.html', context)


from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import timedelta
from .models import ActivityLog

@login_required
def activity_log_view(request):
    """View activity logs with filtering"""
    activities = ActivityLog.objects.all()
    
    # Filters
    user_filter = request.GET.get('user', '')
    action_filter = request.GET.get('action', '')
    date_filter = request.GET.get('date_range', '')
    
    if user_filter:
        activities = activities.filter(user__username__icontains=user_filter)
    
    if action_filter:
        activities = activities.filter(action=action_filter)
    
    if date_filter:
        if date_filter == 'today':
            activities = activities.filter(timestamp__date=timezone.now().date())
        elif date_filter == 'week':
            one_week_ago = timezone.now() - timedelta(days=7)
            activities = activities.filter(timestamp__gte=one_week_ago)
        elif date_filter == 'month':
            one_month_ago = timezone.now() - timedelta(days=30)
            activities = activities.filter(timestamp__gte=one_month_ago)
    
    # Compute action type counts for chart
    action_counts = {}
    for action, _ in ActivityLog.ACTION_TYPES:
        count = activities.filter(action=action).count()
        action_counts[action] = count
    
    # Pagination
    paginator = Paginator(activities, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Log this view access
    log_activity(
        user=request.user,
        action='view',
        description="Accessed activity log",
        request=request
    )
    
    context = {
        'page_obj': page_obj,
        'action_types': ActivityLog.ACTION_TYPES,
        'action_counts': action_counts,
        'selected_user': user_filter,
        'selected_action': action_filter,
        'selected_date_range': date_filter,
    }
    return render(request, 'activity_log.html', context)


@login_required
def user_sessions_view(request):
    """View active and historical user sessions"""
    active_sessions = UserSession.objects.filter(is_active=True)
    recent_sessions = UserSession.objects.filter(is_active=False).order_by('-logout_time')[:50]
    
    log_activity(
        user=request.user,
        action='view',
        description="Accessed user sessions view",
        request=request
    )
    
    context = {
        'active_sessions': active_sessions,
        'recent_sessions': recent_sessions,
    }
    return render(request, 'user_sessions.html', context)


@login_required
def system_health_view(request):
    """System health and performance dashboard"""
    today = timezone.now().date()
    
    # Get today's metrics
    today_metrics, _ = SystemMetrics.objects.get_or_create(date=today)
    
    # Recent activity summary
    activity_summary = {}
    for action_type, _ in ActivityLog.ACTION_TYPES:
        count = ActivityLog.objects.filter(
            timestamp__date=today,
            action=action_type
        ).count()
        activity_summary[action_type] = count
    
    # Active users today
    active_users_today = ActivityLog.objects.filter(
        timestamp__date=today
    ).values('user__username').distinct().count()
    
    # Reports by status
    open_reports = CybercrimeReport.objects.filter(is_resolved=False).count()
    resolved_reports = CybercrimeReport.objects.filter(is_resolved=True).count()
    
    # Average resolution time
    resolved_with_times = CybercrimeReport.objects.filter(
        is_resolved=True,
        resolution_date__isnull=False
    )
    
    total_resolution_time = timedelta()
    count = 0
    for report in resolved_with_times:
        resolution_time = report.resolution_date - report.submitted_at
        total_resolution_time += resolution_time
        count += 1
    
    avg_resolution_time = total_resolution_time / count if count > 0 else timedelta()
    
    log_activity(
        user=request.user,
        action='view',
        description="Accessed system health dashboard",
        request=request
    )
    
    context = {
        'today_metrics': today_metrics,
        'activity_summary': activity_summary,
        'active_users_today': active_users_today,
        'open_reports': open_reports,
        'resolved_reports': resolved_reports,
        'avg_resolution_hours': avg_resolution_time.total_seconds() / 3600,
    }
    return render(request, 'system_health.html', context)